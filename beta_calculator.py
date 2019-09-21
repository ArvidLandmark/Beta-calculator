import requests
from bs4 import BeautifulSoup
import openpyxl as xl


# How do I use a variable as argument in this function?
def paste_excel():
    sheet = wb_new["Sheet1"]
    base_set = ticker.upper()
    cell = sheet.cell(1, 1)
    cell.value = base_set
    for i in range(count_dates):
        for ij in range(3):
            try:
                paste_value = data_lists[ij][i]
                cell = sheet.cell(3 + i, 1 + ij*4)
                cell.value = str(value_checker(paste_value[0]))
                cell = sheet.cell(3 + i, 2 + ij*4)
                cell.value = str(value_checker(paste_value[1]))
            except:
                pass
    wb_new.save(loc_save)


def value_checker(value):
    if "." in value and "," in value:
        value = str(value).replace(",", "")
        value = str(value).replace(".", ",")
    else:
        if "." in value:
            value = str(value).replace(".", ",")
    return value


def historical_prices():
    for io in range(3):
        result = requests.get(f"https://finance.yahoo.com/quote/{ticker_lists[io]}/history?period1=345400200&period2=1568217600&interval=1mo&filter=history&frequency=1mo")
        soup = BeautifulSoup(result.text, 'html.parser')
        if io == 0:
            count = 0
            for ik in range(100):
                try:
                    horse = soup.findAll('tr', {'class': 'BdT Bdc($seperatorColor) Ta(end) Fz(s) Whs(nw)'})[ik + 1].findAll("td", {"class": "Py(10px) Pstart(10px)"})[0].text
                    count += 1
                except:
                    pass
            print(count)
        for i in range(count):
            try:
                tag_date = soup.findAll('tr', {'class': 'BdT Bdc($seperatorColor) Ta(end) Fz(s) Whs(nw)'})[i+1].findAll("td", {"class": "Py(10px) Ta(start) Pend(10px)"})[0].text
                tag_close = soup.findAll('tr', {'class': 'BdT Bdc($seperatorColor) Ta(end) Fz(s) Whs(nw)'})[i+1].findAll("td", {"class": "Py(10px) Pstart(10px)"})[4].text
                print(f"{tag_date}, Close: {tag_close}")
                data_lists[io].append([tag_date, tag_close])
            except:
                continue
    return count


def url_creator(ticker_symbol):
    url = f"https://finance.yahoo.com/quote/{ticker_symbol}?p={ticker_symbol}&.tsrc=fin-srch"
    return url


########################################################################################################################
# Base location
p_loc_excel = r"C:\Users\arvid\OneDrive\Skrivbord\Python\Beta\beta_base.xlsm"

# Asks the user if other location is preferable
"""
user_answer = input(f"Is [{p_loc_excel}] the correct location of your beta_base file? [Y/N]")
if user_answer.upper() == "Y":
    loc_excel = p_loc_excel
else:
    loc_excel = input("Enter your base_file location [including file name.xlsm]: ")
"""
user_answer = input("Enter the location of beta_base: ")
loc_excel = user_answer + "\\beta_base.xlsm"

ticker = input("Enter a ticker: ")
print(f"File will be saved as {loc_excel[:-5]}_{ticker.upper()}.xlsm")

user_s = input("Would you like to use S&P500 as the market? [Y/N] ")
if user_s.upper() == "Y":
    stock_exchange_ticker = "%5EGSPC"
else:
    stock_exchange_ticker = "%5E" + input("Enter the ticker for the stock exchange: [Only text]")
    print(stock_exchange_ticker)

# Loads workbook and creates the necessary lists
wb = xl.load_workbook(loc_excel, read_only=False, keep_vba=True)
wb_new = wb
loc_save = loc_excel[:-5] + "_" + ticker.upper() + ".xlsm"
ticker_lists = [ticker, stock_exchange_ticker, "%5ETNX"]
stock_list, sp_list, rf_list = [], [], []
data_lists = [stock_list, sp_list, rf_list]

# Fetches the data
count_dates = historical_prices()

# Enter the data into workbook
paste_excel()
print(stock_list, "\n", sp_list, "\n", rf_list)

# Beta	=SLOPE(D4:D100;H4:H100)
# Adjusted beta	=N6*2/3+1/3
