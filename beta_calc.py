import requests
from bs4 import BeautifulSoup
import openpyxl as xl
import time

# How do I use a variable as argument in this function?
def paste_excel():
    sheet = wb_new["Sheet1"]
    base_set = ticker.upper()
    cell = sheet.cell(1, 1)
    cell.value = base_set
    market_set = stock_exchange_ticker[3:].upper()
    cell = sheet.cell(1, 6)
    cell.value = market_set
    for i in range(count_dates):
        for ij in range(3):
            try:
                paste_value = data_lists[ij][i]
                cell = sheet.cell(3 + i, 1 + ij*4)
                cell.value = str(value_checker(paste_value[0]))
                cell = sheet.cell(3 + i, 2 + ij*4)
                if ticker_lists[ij] == "%5ETNX":
                    temp_float = str(round(float(paste_value[1]) / 100, 5)).replace(".", ",")
                    cell.value = temp_float
                else:
                    cell.value = str(value_checker(paste_value[1]))
            except:
                pass
    wb_new.save(loc_save)


"""
                if ticker_lists[ij] == "%5ETNX":
                    temp_float = str(float(paste_value[1]) / 100).replace(".", ",")
                    cell.value = temp_float
                else:
                    cell.value = str(value_checker(paste_value[1]))
"""

def value_checker(value):
    if "." in value and "," in value:
        value = str(value).replace(",", "")
        value = str(value).replace(".", ",")
    else:
        if "." in value:
            value = str(value).replace(".", ",")
    return value


def historical_prices():
    for io in range(3):
        result = requests.get(f"https://finance.yahoo.com/quote/{ticker_lists[io]}/history?period1=345400200&period2=1568217600&interval=1mo&filter=history&frequency=1mo")
        soup = BeautifulSoup(result.text, 'html.parser')
        if io == 0:
            count = 0
            for ik in range(100):
                try:
                    horse = soup.findAll('tr', {'class': 'BdT Bdc($seperatorColor) Ta(end) Fz(s) Whs(nw)'})[ik + 1].findAll("td", {"class": "Py(10px) Pstart(10px)"})[0].text
                    count += 1
                except:
                    pass
        for i in range(count):
            try:
                tag_date = soup.findAll('tr', {'class': 'BdT Bdc($seperatorColor) Ta(end) Fz(s) Whs(nw)'})[i+1].findAll("td", {"class": "Py(10px) Ta(start) Pend(10px)"})[0].text
                tag_close = soup.findAll('tr', {'class': 'BdT Bdc($seperatorColor) Ta(end) Fz(s) Whs(nw)'})[i+1].findAll("td", {"class": "Py(10px) Pstart(10px)"})[4].text
                data_lists[io].append([tag_date, tag_close])
            except:
                continue
    return count


def url_creator(ticker_symbol):
    url = f"https://finance.yahoo.com/quote/{ticker_symbol}?p={ticker_symbol}&.tsrc=fin-srch"
    return url

#################

string = input("Enter stock tickers [separated with a ,: ")           # betas to create # "A,AAL,AAP,AAPL,ABBV,ABC,ABMD,ABT,ACN,ADBE,ADI,ADM,ADP,ADS,ADSK,AEE,AEP,AES,AFL,AGN,AIG,AIV,AIZ,AJG,AKAM,ALB,ALGN,ALK,ALL,ALLE,ALXN,AMAT,AMCR,AMD,AME,AMG,AMGN,AMP,AMT,AMZN,ANET,ANSS,ANTM,AON,AOS,APA,APD,APH,APTV,ARE,ARNC,ATO,ATVI,AVB,AVGO,AVY,AWK,AXP,AZO,BA,BAC,BAX,BBT,BBY,BDX,BEN,BF.B,BHGE,BIIB,BK,BKNG,BLK,BLL,BMY,BR,BRK.B,BSX,BWA,BXP,C,CAG,CAH,CAT,CB,CBOE,CBRE,CBS,CCI,CCL,CDNS,CE,CELG,CERN,CF,CFG,CHD,CHRW,CHTR,CI,CINF,CL,CLX,CMA,CMCSA,CME,CMG,CMI,CMS,CNC,CNP,COF,COG,COO,COP,COST,COTY,CPB,CPRI,CPRT,CRM,CSCO,CSX,CTAS,CTL,CTSH,CTVA,CTXS,CVS,CVX,CXO,D,DAL,DD,DE,DFS,DG,DGX,DHI,DHR,DIS,DISCA,DISCK,DISH,DLR,DLTR,DOV,DOW,DRE,DRI,DTE,DUK,DVA,DVN,DXC,EA,EBAY,ECL,ED,EFX,EIX,EL,EMN,EMR,EOG,EQIX,EQR,ES,ESS,ETFC,ETN,ETR,EVRG,EW,EXC,EXPD,EXPE,EXR,F,FANG,FAST,FB,FBHS,FCX,FDX,FE,FFIV,FIS,FISV,FITB,FLIR,FLS,FLT,FMC,FOX,FOXA,FRC,FRT,FTI,FTNT,FTV,GD,GE,GILD,GIS,GL,GLW,GM,GOOG,GOOGL,GPC,GPN,GPS,GRMN,GS,GWW,HAL,HAS,HBAN,HBI,HCA,HCP,HD,HES,HFC,HIG,HII,HLT,HOG,HOLX,HON,HP,HPE,HPQ,HRB,HRL,HSIC,HST,HSY,HUM,IBM,ICE,IDXX,IEX,IFF,ILMN,INCY,INFO,INTC,INTU,IP,IPG,IPGP,IQV,IR,IRM,ISRG,IT,ITW,IVZ,JBHT,JCI,JEC,JEF,JKHY,JNJ,JNPR,JPM,JWN,K,KEY,KEYS,KHC,KIM,KLAC,KMB,KMI,KMX,KO,KR,KSS,KSU,L,LB,LDOS,LEG,LEN,LH,LHX,LIN,LKQ,LLY,LMT,LNC,LNT,LOW,LRCX,LUV,LW,LYB,M,MA,MAA,MAC,MAR,MAS,MCD,MCHP,MCK,MCO,MDLZ,MDT,MET,MGM,MHK,MKC,MKTX,MLM,MMC,MMM,MNST,MO,MOS,MPC,MRK,MRO,MS,MSCI,MSFT,MSI,MTB,MTD,MU,MXIM,MYL,NBL,NCLH,NDAQ,NEE,NEM,NFLX,NI,NKE,NKTR,NLSN,NOC,NOV,NRG,NSC,NTAP,NTRS,NUE,NVDA,NWL,NWS,NWSA,O,OKE,OMC,ORCL,ORLY,OXY,PAYX,PBCT,PCAR,PEG,PEP,PFE,PFG,PG,PGR,PH,PHM,PKG,PKI,PLD,PM,PNC,PNR,PNW,PPG,PPL,PRGO,PRU,PSA,PSX,PVH,PWR,PXD,PYPL,QCOM,QRVO,RCL,RE,REG,REGN,RF,RHI,RJF,RL,RMD,ROK,ROL,ROP,ROST,RSG,RTN,SBAC,SBUX,SCHW,SEE,SHW,SIVB,SJM,SLB,SLG,SNA,SNPS,SO,SPG,SPGI,SRE,STI,STT,STX,STZ,SWK,SWKS,SYF,SYK,SYMC,SYY,T,TAP,TDG,TEL,TFX,TGT,TIF,TJX,TMO,TMUS,TPR,TRIP,TROW,TRV,TSCO,TSN,TSS,TTWO,TWTR,TXN,TXT,UA,UAA,UAL,UDR,UHS,ULTA,UNH,UNM,UNP,UPS,URI,USB,UTX,V,VAR,VFC,VIAB,VLO,VMC,VNO,VRSK,VRSN,VRTX,VTR,VZ,WAB,WAT,WBA,WCG,WDC,WEC,WELL,WFC,WHR,WLTW,WM,WMB,WMT,WRK,WU,WY,WYNN,XEC,XEL,XLNX,XOM,XRAY,XRX,XYL,YUM,ZBH,ZION,ZTS"

all_tickers = string.split(",")


########################################################################################################################
loc_excel = r"C:\Users\arvid\OneDrive\Skrivbord\Python\Beta\beta_base.xlsm"
print(loc_excel)

user_s = input("Would you like to use S&P500 as the market? [Y/N] ")
if user_s.upper() == "Y":
    stock_exchange_ticker = "%5EGSPC"
else:
    stock_exchange_ticker = "%5E" + input("Enter the ticker for the stock exchange: [Only text]")
    print(stock_exchange_ticker)

print("Loading..")

for it in range(len(all_tickers)):
    ticker = all_tickers[it]
    wb = xl.load_workbook(loc_excel, read_only=False, keep_vba=True)
    wb_new = wb
    loc_save = loc_excel[:-5] + "_" + ticker + ".xlsm"
    ticker_lists = [ticker, stock_exchange_ticker, "%5ETNX"]
    stock_list, sp_list, rf_list = [], [], []
    data_lists = [stock_list, sp_list, rf_list]
    count_dates = historical_prices()
    paste_excel()
    print(f"[{it+1}/{len(all_tickers)}] Created: {all_tickers[it].upper()}")
    print(f"Saved as: {loc_save}")
    time.sleep(0.5)

print("All tickers searched.")





